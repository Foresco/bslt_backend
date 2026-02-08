# Классы источников данных
import xlrd  # Библиотека для работы с файлами Excel xls
import openpyxl  # Библиотека для работы с файлами Excel xlsx
import lxml.etree as lxmllib # Библиотека для работы с xml
import xmltodict
import datetime  # Библиотека для работы с датами


class SourceExcel:

    def __init__(self, **kwargs):
        self.file_name = kwargs['file_name']
        self.sheet_name = kwargs['sheet_name']
        self.max_col_num = 0  # Количество столбцов в таблице
        self.max_row_num = 0  # Количество строк в таблице
        self.header_columns = dict()

    def open(self):
        # Открытие файла
        try:
            self.work_book = xlrd.open_workbook(self.file_name)
        except:
            print('Ошибка открытия файла', self.file_name)
            self.work_book = ''
            raise
        # Открытие листа
        try:
            self.sheet = self.work_book.sheet_by_name(self.sheet_name)  # Открытие листа с данными
        except:
            print('Ошибка открытия листа', self.sheet_name)
            self.sheet_name = ''
            raise

        self.headers_get()  # Получение заголовков листа

    def headers_get(self):
        """ Получение массива с заголовками """
        row_num = 0
        col_num = 0
        header_columns = dict()
        # Определение размера данных на листе
        self.max_col_num = self.sheet.ncols - 1  # Определение количества столбцов в таблице
        self.max_row_num = self.sheet.nrows      # Определение количества строк в таблице

        while not col_num > self.max_col_num:  # Пока не вышли за рамки данных на листе
            cell = self.sheet.cell(row_num, col_num)
            if cell.value > '':  # Только столбцы, имеющие название
                self.header_columns[col_num] = cell.value
            col_num += 1

    def names_translate(self, row):
        """Замена числовых номеров полей на текстовые названия"""
        new_row = dict()
        i = 0
        for cell in row:
            new_row[self.header_columns[i]] = cell.value
            i += 1
        return new_row

    def row_get(self):
        """Итератор по строкам таблицы"""
        for row_num in range(1, self.max_row_num):
            yield self.names_translate(self.sheet.row(row_num))

    def float_to_date_translate(self, str_value):
        """Предобразование числовой даты в традиционную"""
        # print(self.work_book.datemode)
        return datetime.datetime(*xlrd.xldate_as_tuple(str_value, self.work_book.datemode)).strftime("%d.%m.%Y")

    def close(self):
        # Закрытие Excel файла
        self.work_book.release_resources()


class SourceExcelx(SourceExcel):
    """Вариант для файлов xlsx"""

    def open(self):
        # Открытие файла
        try:
            self.work_book = openpyxl.load_workbook(self.file_name)
        except:
            print('Ошибка открытия файла', self.file_name)
            self.work_book = ''
            raise
        # Открытие листа
        try:
            self.sheet = self.work_book.get_sheet_by_name(self.sheet_name)  # Открытие листа с данными
        except:
            print('Ошибка открытия листа', self.sheet_name)
            self.sheet_name = ''
            raise

        self.headers_get()  # Получение заголовков листа

    def headers_get(self):
        """ Получение массива с заголовками """
        row_num = 1
        col_num = 1
        # Определение размера данных на листе
        self.max_col_num = self.sheet.max_column  # Определение количества столбцов в таблице
        self.max_row_num = self.sheet.max_row      # Определение количества строк в таблице

        while not col_num > self.max_col_num:  # Пока не вышли за рамки данных на листе
            cell = self.sheet.cell(row_num, col_num)
            if cell.value:  # Только столбцы, имеющие название
                self.header_columns[col_num] = cell.value
            col_num += 1
    
    def names_translate(self, row):
        """Замена числовых номеров полей на текстовые названия"""
        new_row = dict()
        i = 1
        for cell in row:
            new_row[self.header_columns[i]] = cell.value
            i += 1
        return new_row

    def row_get(self):
        """Итератор по строкам таблицы"""
        for row in self.sheet.iter_rows():
            yield self.names_translate(row)
    
    def close(self):
        # Закрытие Excel файла
        self.work_book.close()        


class SourceXml:

    def __init__(self, **kwargs):
        self.file_name = kwargs['file_name']

        # Открытие файла
        try:
            self.tree = lxmllib.parse(self.file_name)
            self.root_node = self.tree.getroot()
        except:
            print('Ошибка открытия xml-файла с данными ', self.file_name)
            raise

    def row_get(self):
        """Итератор по веткам xml"""
        for xml_node in self.root_node:
            yield xmltodict.parse(xml_node)