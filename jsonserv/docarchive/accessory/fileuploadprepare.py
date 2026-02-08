# Подготовка данных для загрузки информации об архивных документах из таблицы
import openpyxl
from parse import parse

from jsonserv.core.models import fn_head_key
from jsonserv.pdm.models import Designer, PartFormat, PartObject, PartObjectFormat, Stage
from jsonserv.docarchive.models import ArcDocument, ArcDocumentObject, CodePrefix, DocumentType, FileUpload, UploadArcdoc


def replace_double_space(srcstr):
    # Чистка лишних пробелов
    result = srcstr
    while result.find('  ') > 0:
        result = result.replace('  ', ' ')
    return result


class FileUploadImport:
    # Колонки таблицы
    table_columns = (
        "Проект",
        "Дата",
        "Обозначение",
        "Наименование",
        "Количество листов в документе (формат)",
        "Кто передал"
    )

    def __init__(self, src_file, file_name, stage, prefix, crtd_sess_id):
        self.src_file = src_file  # Расположение загружаемого файла
        self.file_name = file_name  # Имя файла
        self.upoad_id = None  # Идентификатор созданной загрузки
        self.stage = Stage.objects.get(pk=stage) if stage else None
        self.prefix = CodePrefix.objects.get(pk=prefix) if prefix else None

        self.crtd_sess_id = crtd_sess_id
        self.col_numbers = {
            'Инв.№': 1  # Инвентарный номер всегда первая колонка
        }  # Номера колонок
        self.header_row = 0
        self.rows_data = []  # Собранные для загрузки строки
        self.part_objects = dict()  # Список объектов
        # Список доступных форматов
        self.sheet_formats = dict(
            map(lambda x: (x.list_value, x), PartFormat.objects.all()))
        self.report = ''  # Сообщения о результатах

    @staticmethod
    def caption_prepare(caption):
        result = caption.replace('СМ', 'CM')
        result = result.replace('С.', 'C.')
        return result.strip()

    @staticmethod
    def name_prepare(name_value):
        # Строки, которые нужно исключить из наименования
        str_exclude = ('Спецификация', 'Сборочный чертеж', 'Ведомость покупных изделий')
        result = name_value
        # Удаление лишних слов
        for strng in str_exclude:
            result = result.replace(strng, '')

        result = result.strip()
        result = replace_double_space(result)  # Чистка лишних пробелов
        return result

    def parse_format(self, format_str):
        """Разбор строки с форматами"""
        if not format_str:
            return {}
        # Обрезка пробелов по краям
        format_str = format_str.strip()
        format_str = format_str.replace('   ', ',')  # Вставка запятой как разделителя вместо трех пробелов
        # Удаление пробелов
        format_str = format_str.replace(' ', '')
        # Чистка двойных запятых
        while format_str.find(',,') > 0:
            format_str = format_str.replace(',,', ',')

        raw_formats = format_str.split(',')
        result = list()
        list_count = 0  # Счетчик листов

        for raw_format in raw_formats:
            list_quantity = 1
            # Замена русской x латинской
            raw_format = raw_format.replace('х', 'x')
            # Замена русской A латинской
            raw_format = raw_format.replace('А', 'A')

            if raw_format.find('х') > 0 and raw_format.find('(') > 0:
                # print(1)
                list_quantity, format_number = parse('{:d}({})', raw_format)  # '%d(%4[^х]х%d)'
            elif raw_format.find('х') > 0 and not raw_format.find('(') > 0:
                # print(2)
                format_number, format_quantity = parse('{}х{:d}', raw_format)  # '%4[^х]х%d'
            elif not raw_format.find('х') > 0 and raw_format.find('(') > 0:
                # print(3)
                list_quantity, format_number = parse('{:d}({})', raw_format)  # '%d(%4[^)])'
            elif not raw_format.find('х') > 0 and not raw_format.find('(') > 0:
                # print(4)
                format_number = raw_format

            if format_number not in self.sheet_formats:
                return f'Ошибка при разборе формата - неизвестный формат [{format_number}]'

            if list_quantity < 1:
                return f'Ошибка при разборе формата - не верное количество листов [{raw_format}]'

            result.append({
                'list_quantity': list_quantity,
                'format_id': self.sheet_formats[format_number]
            })
            list_count += list_quantity  # Считаем листы

        return result, list_count

    def get_header(self, sheet):
        """Разбор заголовка переданной таблицы"""
        row = 1
        while row < 20:
            row += 1
            cell_value = sheet.cell(column=1, row=row).value
            if 'Инв.№' == cell_value:
                self.header_row = row + 1  # Там есть еще строка с номерами столбцов
                column = 1
                while sheet.cell(column=column, row=row).value:
                    cell_value = sheet.cell(column=column, row=row).value
                    if cell_value in self.table_columns:
                        self.col_numbers[cell_value] = column
                    column += 1
                if "Обозначение" not in self.col_numbers:
                    return 'Ошибка при разборе файла - отсутствует столбец [Обозначение]'
                else:
                    return ''
        return 'Ошибка при разборе файла - заголовок таблицы не найден'

    def parse_table(self, sheet):
        """Разбор основного тела таблицы"""
        row = self.header_row + 1
        while sheet.cell(column=self.col_numbers['Обозначение'], row=row).value:
            date = sheet.cell(column=self.col_numbers['Дата'], row=row).value
            project = sheet.cell(column=self.col_numbers['Проект'], row=row).value if (
                    'Проект' in self.col_numbers) else self.prefix
            inv_num = sheet.cell(column=self.col_numbers['Инв.№'], row=row).value
            caption = self.caption_prepare(sheet.cell(column=self.col_numbers['Обозначение'], row=row).value)
            name = self.name_prepare(sheet.cell(column=self.col_numbers['Наименование'], row=row).value) if (
                    'Наименование' in self.col_numbers) else None
            formats = sheet.cell(column=self.col_numbers['Количество листов в документе (формат)'], row=row).value if (
                    'Количество листов в документе (формат)' in self.col_numbers) else None
            who = sheet.cell(column=self.col_numbers['Кто передал'], row=row).value if (
                    'Кто передал' in self.col_numbers) else None

            if name and len(name) > 250:
                return f'Ошибка при разборе таблицы - длина значения в столбце "Наименование" больше 250. Строка {row}'
            if not inv_num:
                return f'Ошибка при разборе таблицы - в столбце "Инв.№" некорректное значение. Строка {row}'
            if not date:
                return f'Ошибка при разборе таблицы - в столбце "Дата" некорректное значение. Строка {row}'
            if not who:
                return f'Ошибка при разборе таблицы - в столбце "Кто передал" отсутствует значение. Строка {row}'
            if not caption:
                return f'Ошибка при разборе таблицы - в столбце "Обозначение" отсутствует значение. Строка {row}'

            format_d, list_count = self.parse_format(formats)  # Разбор строки форматов
            if type(format_d) is str:  # Это сообщение об ошибке
                return format_d

            # Добавляем разобранную строку в список
            self.rows_data.append({
                'date': date,
                'project': project,
                'inv_num': inv_num,
                'caption': caption,
                'name': name,
                'format': format_d,
                'list_count': list_count,
                'who': who,
                'first': caption  # Привязываем к объекту с таким же обозначением
            })

            row += 1

        return ''

    def check_rows_data(self):
        projects = list()
        designers = list()

        for row in self.rows_data:
            project = row['project']
            if project and project not in projects:
                projects.append(project)

            who = row['who'].strip()
            if who and who not in designers:
                designers.append(who)

            first = row['first'].strip()
            if first:
                # Стадию не добавляем, так как отказались от составного ключа для ДСЕ
                if (1 == 2) and self.stage:
                    head_key = fn_head_key(first, self.stage.code)
                    mess = f"{first} ({self.stage})"
                else:
                    head_key = fn_head_key(first)
                    mess = f"{first}"
                # Ищем по ключу, так как возможна путаница с рус/лат
                part_object = PartObject.objects.filter(head_key=head_key).first()
                if part_object:
                    self.part_objects[first] = part_object
                else:
                    self.report += f"Объект состава {mess} не найден.\n"

        # Проверка проектов
        for project in projects:
            if not self.prefix:
                ps = CodePrefix.objects.filter(prefix_code=project).first()
                if not ps:
                    return f'Ошибка при разборе таблицы - проект [{project}] не найден'

        # Проверка разработчиков
        for designer in designers:
            ds = Designer.objects.filter(designer=designer)
            if not ds:
                return f'Ошибка при разборе таблицы - разработчик [{designer}] не найден'
        return 0

    def import_rows_data(self):
        """Создание данных на основе ранее подготовленных строк"""
        # Тип Конструкторский документ в качестве значения по умолчанию
        doc_type_kd = DocumentType.objects.get(pk=2)

        for row in self.rows_data:
            doc = 0
            params = dict(
                prefix=row['project'],
                document_num=row['inv_num'],
                code=row['caption'],
                document_name=row['name'],
                reg_date=row['date'],
                parent=self.stage,
                doc_type=doc_type_kd,
		list_count=row['list_count'],
                crtd_sess_id=self.crtd_sess_id
            )

            # Создание архивного документа
            doc, result = ArcDocument.get_or_create_item(params)
            if not result:
                return f"Архивный документ с обозначением [{row['caption']}] уже зарегистрирован ранее"
            self.report += f"{row['caption']} {row['name']} {row['date']} успешно добавлен.\n"

            # Связь архивного документа и загрузки
            link_id, result = UploadArcdoc.get_or_create_item(
                dict(file_upload=self.file_upload, arc_doc=doc, crtd_sess_id=self.crtd_sess_id))

            # Добавление форматов
            for frmt in row['format']:
                link_id, result = PartObjectFormat.get_or_create_item(
                    dict(part_object=doc, list_quantity=frmt['list_quantity'], format=frmt['format_id'],
                         crtd_sess_id=self.crtd_sess_id))

            # Добавление применения если указано и есть соответствующий объект
            if row['first'] in self.part_objects:
                link_id, result = ArcDocumentObject.get_or_create_item(
                    dict(parent=doc, child=self.part_objects[row['first']],
                         crtd_sess_id=self.crtd_sess_id))
                self.report += f"{row['first']} успешно связан с архивным документом.\n"
            else:
                self.report += f"{row['first']} не был связан с архивным документом!\n"

    def import_table(self):
        """Загрузка информации из таблицы"""
        # Открываем переданный файл
        workbook = openpyxl.load_workbook(self.src_file)

        # Переходим на активную страницу
        worksheet = workbook.active

        # Разбираем заголовок
        err = self.get_header(worksheet)
        if err:
            return err
        # Разбор таблицы на массивы
        err = self.parse_table(worksheet)
        if err:
            return err
        # Проверка полученных данных
        err = self.check_rows_data()
        if err:
            return err
        # Создание загрузки
        self.file_upload, result = FileUpload.get_or_create_item(
            dict(file_name=self.file_name, crtd_sess_id=self.crtd_sess_id))
        # Загрузка обработанных данных в базу
        err = self.import_rows_data()
        if err:
            return err
